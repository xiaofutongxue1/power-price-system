# -*- coding: utf-8 -*-
# pages/07_模板数据集生成.py

import streamlit as st
from io import BytesIO
import pandas as pd
from openpyxl.styles import Font, Alignment

# ================================
# 常量配置
# ================================
STATION_TYPE_DEFAULT = "对外开放站点"
OPEN_RULE_DEFAULT = "全终端全时段对外开放"
DEFAULT_STRATEGY_TEXT = (
    "在投资回收测算的目标服务费基础上结合周边竞品制定服务费，"
    "站点上线后根据实际运营情况对服务费进行灵活调整（调整幅度±20%）"
)

# 模板列名（单层表头）
ALL_COLS = [
    "序号",
    "站点名称",
    "站点编号",
    "站点类型",
    "开放规则",
    "定价策略-总策略",
    "定价策略-基础电费",
    "定价策略-服务费",
    "定价策略-超时占位费",
    "定价策略-停车费",
    "价格生效时间",
    "本次生效价格-电费",
    "本次生效价格-服务费",
    "本次生效价格-总电价",
    "本次生效价格-超时占位费",
    "本次生效价格-停车费",
    "竞品价格",
]

# ================================
# 页面标题
# ================================
st.markdown("""
<div class='main-header'>
📑 价格模板数据集生成
</div>
""", unsafe_allow_html=True)

st.markdown("""
<div class='sub-header'>
本页面在既有的电费 / 服务费 / 总价计算结果基础上，生成
<strong>「序号 + 站点信息 + 定价策略 + 本次生效价格」</strong> 的完整价格模板 Excel。
</div>
""", unsafe_allow_html=True)

# ================================
# 1. 上传三张“结构表”
# ================================
st.markdown("<div class='card'>", unsafe_allow_html=True)
st.markdown("""
<div class='card-title'>
  <div class='icon-circle'>📄</div>
  ① 上传电费 / 服务费结构表 & 当前服务费均价表
</div>

- <b>电费价格时段表</b>：需要包含至少 <code>序号</code>、<code>站点编号</code>、<code>供电规则</code> 等列；  
- <b>服务费价格时段表</b>：需要包含至少 <code>站点全称</code>、<code>站点编号</code>、<code>站点名称</code>、<code>目标服务费</code> 等列；  
- <b>当前服务费均价表</b>：需要包含至少 <code>站点名称</code>、<code>当前服务费均价</code> 等列。
""", unsafe_allow_html=True)

file_elec_struct = st.file_uploader(
    "电费价格时段表（Excel）",
    type=["xlsx"],
    key="file_elec_struct",
)

file_serv_struct = st.file_uploader(
    "服务费价格时段表（Excel）",
    type=["xlsx"],
    key="file_serv_struct",
)

file_serv_avg = st.file_uploader(
    "当前服务费均价表（Excel）",
    type=["xlsx"],
    key="file_serv_avg",
)

st.markdown("</div>", unsafe_allow_html=True)

# ================================
# 2. 统一策略文本 & 生效时间
# ================================
st.markdown("<div class='card'>", unsafe_allow_html=True)
st.markdown("""
<div class='card-title'>
  <div class='icon-circle'>📝</div>
  ② 填写统一的策略 & 生效时间
</div>
""", unsafe_allow_html=True)

strategy_text = st.text_area(
    "统一填写「定价策略-总策略」文本（所有行默认相同，可导出后再逐行微调）：",
    value=DEFAULT_STRATEGY_TEXT,
    height=80,
)

effective_time = st.text_input(
    "统一填写「价格生效时间」（例如：2025-01-01 或 2025/01/01 00:00:00）：",
    value="",
)

st.markdown("</div>", unsafe_allow_html=True)

# ================================
# 3. 电费 / 服务费 / 总价 结果来源
# ================================
st.markdown("<div class='card'>", unsafe_allow_html=True)
st.markdown("""
<div class='card-title'>
  <div class='icon-circle'>⚡</div>
  ③ 电费 / 服务费 / 总价 结果来源
</div>

优先使用本系统前几页（Page3 / Page4-5 / Page6）已经计算好的结果；
如果 session 中未检测到，则需要你手动上传。
""", unsafe_allow_html=True)

# Page3：电费结果
power_df_state = st.session_state.get("station_fee")
# Page4：服务费结果原始
service_df_state = st.session_state.get("service_price_raw")
# Page5：服务费矫正结果映射
corrected_map = st.session_state.get("service_price_corrected", {})
# Page6：总价结果
total_df_state = st.session_state.get("total_price_result")

need_power_upload = not isinstance(power_df_state, pd.DataFrame) or power_df_state.empty
need_serv_upload = not isinstance(service_df_state, pd.DataFrame) or service_df_state.empty
need_total_upload = not isinstance(total_df_state, pd.DataFrame) or total_df_state.empty

col_p, col_s, col_t = st.columns(3)

# ---- 电费结果 ----
with col_p:
    if not need_power_upload:
        st.success("已检测到 Page3 的电费结果，可直接使用。")
        power_df = power_df_state.copy()
        power_file_upload = None
    else:
        st.warning("未检测到 Page3 的电费结果，请上传电费结果 Excel（含『站点名称』『电费』列）。")
        power_file_upload = st.file_uploader(
            "上传电费结果表",
            type=["xlsx"],
            key="upload_power_result",
        )
        power_df = None

# ---- 服务费结果 ----
with col_s:
    if not need_serv_upload:
        st.success("已检测到 Page4/5 的服务费结果，可直接使用（优先使用 Page5 矫正后数据）。")

        raw = service_df_state.copy()
        station_list = raw["站点名称"].unique().tolist()

        if isinstance(corrected_map, dict) and corrected_map:
            rows = []
            for name in station_list:
                if name in corrected_map:
                    segs = corrected_map[name]
                    txt = "\n".join(
                        [f"{s['start']} - {s['end']} {s['price']}元/度" for s in segs]
                    )
                else:
                    txt = str(raw.loc[raw["站点名称"] == name, "服务费"].values[0])
                rows.append({"站点名称": name, "服务费": txt})
            service_df = pd.DataFrame(rows)
        else:
            service_df = raw[["站点名称", "服务费"]].copy()

        serv_file_upload = None
    else:
        st.warning("未检测到 Page4/5 的服务费结果，请上传服务费结果 Excel（含『站点名称』『服务费』列）。")
        serv_file_upload = st.file_uploader(
            "上传服务费结果表",
            type=["xlsx"],
            key="upload_serv_result",
        )
        service_df = None

# ---- 总价结果 ----
with col_t:
    if not need_total_upload:
        st.success("已检测到 Page6 的总价结果，可直接使用。")
        total_df = total_df_state.copy()
        total_file_upload = None
    else:
        st.warning("未检测到 Page6 的总价结果，请上传总价结果 Excel（含『站点名称』『总价』列）。")
        total_file_upload = st.file_uploader(
            "上传总价结果表",
            type=["xlsx"],
            key="upload_total_result",
        )
        total_df = None

st.markdown("</div>", unsafe_allow_html=True)

# ================================
# 4. 点击生成模板
# ================================
st.markdown("<div class='card'>", unsafe_allow_html=True)
st.markdown("""
<div class='card-title'>
  <div class='icon-circle'>🧮</div>
  ④ 生成价格模板数据集
</div>
""", unsafe_allow_html=True)

if st.button("▶ 生成价格模板数据集", use_container_width=True):

    # -------- 4.1 检查三张结构表 --------
    if file_elec_struct is None or file_serv_struct is None or file_serv_avg is None:
        st.error("❌ 请先上传：电费价格时段表 / 服务费价格时段表 / 当前服务费均价表。")
        st.stop()

    df_elec_struct = pd.read_excel(file_elec_struct)
    df_serv_struct = pd.read_excel(file_serv_struct)
    df_serv_avg = pd.read_excel(file_serv_avg)

    required_elec_cols = {"序号", "站点编号", "供电规则"}
    required_serv_cols = {"站点全称", "站点编号", "站点名称", "目标服务费"}
    required_avg_cols = {"站点名称", "当前服务费均价"}

    if not required_elec_cols.issubset(df_elec_struct.columns):
        st.error(f"❌ 电费价格时段表缺少列：{required_elec_cols - set(df_elec_struct.columns)}")
        st.stop()

    if not required_serv_cols.issubset(df_serv_struct.columns):
        st.error(f"❌ 服务费价格时段表缺少列：{required_serv_cols - set(df_serv_struct.columns)}")
        st.stop()

    if not required_avg_cols.issubset(df_serv_avg.columns):
        st.error(f"❌ 当前服务费均价表缺少列：{required_avg_cols - set(df_serv_avg.columns)}")
        st.stop()

    # -------- 4.2 电费 / 服务费 / 总价结果表处理 --------
    # 电费结果
    if power_df is None and power_file_upload is not None:
        power_df = pd.read_excel(power_file_upload)
    if power_df is None:
        st.error("❌ 仍未获取到电费结果表，请上传或回到 Page3 先计算。")
        st.stop()
    if not {"站点名称", "电费"}.issubset(power_df.columns):
        st.error("❌ 电费结果表必须包含列：『站点名称』『电费』。")
        st.stop()

    # 服务费结果
    if service_df is None and serv_file_upload is not None:
        service_df = pd.read_excel(serv_file_upload)
    if service_df is None:
        st.error("❌ 仍未获取到服务费结果表，请上传或先在 Page4/5 生成。")
        st.stop()
    if not {"站点名称", "服务费"}.issubset(service_df.columns):
        st.error("❌ 服务费结果表必须包含列：『站点名称』『服务费』。")
        st.stop()

    # 总价结果
    if total_df is None and total_file_upload is not None:
        total_df = pd.read_excel(total_file_upload)
    if total_df is None:
        st.error("❌ 仍未获取到总价结果表，请上传或先在 Page6 生成。")
        st.stop()
    if "总价" not in total_df.columns and "总电价" not in total_df.columns:
        st.error("❌ 总价结果表必须包含列：『总价』或『总电价』。")
        st.stop()
    if "总电价" not in total_df.columns and "总价" in total_df.columns:
        total_df = total_df.rename(columns={"总价": "总电价"})

    # -------- 4.3 组装基础表（以服务费结构表为主） --------
    base = df_serv_struct[["站点编号", "站点全称", "站点名称", "目标服务费"]].copy()

    elec_seq = df_elec_struct[["站点编号", "序号", "供电规则"]].copy()
    base = base.merge(elec_seq, on="站点编号", how="left")

    avg_small = df_serv_avg[["站点名称", "当前服务费均价"]].copy()
    base = base.merge(avg_small, on="站点名称", how="left")

    def make_service_strategy(row):
        tgt = row.get("目标服务费")
        cur = row.get("当前服务费均价")
        parts = []
        if pd.notna(tgt):
            parts.append(f"服务费目标均价{tgt}元/度")
        if pd.notna(cur):
            parts.append(f"当前{cur}元/度")
        return "，".join(parts) if parts else ""

    base["策略_服务费"] = base.apply(make_service_strategy, axis=1)

    # 合并电费 / 服务费 / 总价文本（按【站点简称】匹配）
    base = base.merge(
        power_df[["站点名称", "电费"]],
        on="站点名称",
        how="left",
    )

    base = base.merge(
        service_df[["站点名称", "服务费"]],
        on="站点名称",
        how="left",
        suffixes=("", "_服务费结果"),
    )

    base = base.merge(
        total_df[["站点名称", "总电价"]],
        on="站点名称",
        how="left",
    )

    # -------- 4.4 生成最终模板 DataFrame --------
    df_tpl = pd.DataFrame({
        "序号": base["序号"],
        "站点名称": base["站点全称"],      # 最终展示用全称
        "站点编号": base["站点编号"],
        "站点类型": STATION_TYPE_DEFAULT,
        "开放规则": OPEN_RULE_DEFAULT,
        "定价策略-总策略": strategy_text,
        "定价策略-基础电费": base["供电规则"],
        "定价策略-服务费": base["策略_服务费"],
        "定价策略-超时占位费": "本次无变动",
        "定价策略-停车费": "本次无变动",
        "价格生效时间": effective_time,
        "本次生效价格-电费": base["电费"],
        "本次生效价格-服务费": base["服务费"],
        "本次生效价格-总电价": base["总电价"],
        "本次生效价格-超时占位费": "/",
        "本次生效价格-停车费": "/",
        "竞品价格": "/",
    })

    # 列顺序
    df_tpl = df_tpl[ALL_COLS]

    # ① 站点编号强制转为字符串，避免 Excel 截断大整数
    df_tpl["站点编号"] = df_tpl["站点编号"].apply(
        lambda x: "" if pd.isna(x) else str(x)
    )

    st.success(f"✅ 模板数据集生成完成，共 {len(df_tpl)} 行。")
    st.dataframe(df_tpl, use_container_width=True)

    # 提示缺失价格的站点
    missing_price = df_tpl[
        df_tpl["本次生效价格-电费"].isna() |
        df_tpl["本次生效价格-服务费"].isna() |
        df_tpl["本次生效价格-总电价"].isna()
    ]
    if not missing_price.empty:
        st.warning("⚠ 以下站点未完全匹配到电费 / 服务费 / 总电价，请检查名称或结构表：")
        st.dataframe(missing_price[["序号", "站点名称", "站点编号"]])

    # 保存到 session
    st.session_state["price_template_df"] = df_tpl

    # ================== 写 Excel + 样式设置 ==================
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_tpl.to_excel(writer, index=False, sheet_name="价格模板")

        wb = writer.book
        ws = wb["价格模板"]

        # 行列信息
        max_row = ws.max_row
        max_col = ws.max_column

        # 字体
        header_font = Font(name="微软雅黑 Light", size=10, bold=True)
        body_font = Font(name="微软雅黑 Light", size=10)
        price_font_red = Font(name="微软雅黑 Light", size=10, color="FF0000")

        # 对齐：水平居中 + 垂直居中 + 自动换行
        align_center_wrap = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        # 先统一设置对齐 + 基础字体
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = align_center_wrap
                if cell.row == 1:
                    cell.font = header_font
                else:
                    cell.font = body_font

        # 列宽统一设置一下
        for col in ws.columns:
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = 20

        # ② 合并「站点类型 / 开放规则 / 定价策略-总策略」三列（2 行到最后一行）
        if max_row >= 2:
            col_idx_type = df_tpl.columns.get_loc("站点类型") + 1
            col_idx_open = df_tpl.columns.get_loc("开放规则") + 1
            col_idx_strategy = df_tpl.columns.get_loc("定价策略-总策略") + 1

            ws.merge_cells(
                start_row=2, start_column=col_idx_type,
                end_row=max_row, end_column=col_idx_type
            )
            ws.merge_cells(
                start_row=2, start_column=col_idx_open,
                end_row=max_row, end_column=col_idx_open
            )
            ws.merge_cells(
                start_row=2, start_column=col_idx_strategy,
                end_row=max_row, end_column=col_idx_strategy
            )

        # ③ 本次生效价格-电费 / 服务费 / 总电价 三列字体改为红色
        col_idx_price_e = df_tpl.columns.get_loc("本次生效价格-电费") + 1
        col_idx_price_s = df_tpl.columns.get_loc("本次生效价格-服务费") + 1
        col_idx_price_t = df_tpl.columns.get_loc("本次生效价格-总电价") + 1

        for row_idx in range(2, max_row + 1):
            ws.cell(row=row_idx, column=col_idx_price_e).font = price_font_red
            ws.cell(row=row_idx, column=col_idx_price_s).font = price_font_red
            ws.cell(row=row_idx, column=col_idx_price_t).font = price_font_red

    buf.seek(0)

    st.download_button(
        "📥 下载价格模板 Excel",
        data=buf.getvalue(),
        file_name="岚图超充站_价格模板_含策略与价格.xlsx",
        mime=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        use_container_width=True,
    )

st.markdown("</div>", unsafe_allow_html=True)

